"""
استيراد وتصدير Excel فعلي باستخدام openpyxl.
- import_recruits_from_excel : قراءة ملف Excel وإضافة المجندين (مع تخطي المكرر).
- export_recruits_to_excel   : تصدير قائمة المجندين إلى ملف Excel منسّق.
- generate_import_template    : إنشاء قالب استيراد فارغ بالأعمدة الصحيحة.
"""
from __future__ import annotations
import os
from typing import Optional, List, Dict, Any, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from . import services, config

# خريطة: عنوان العمود في ملف Excel -> الحقل في قاعدة البيانات
# تدعم عدة تسميات شائعة لكل حقل
COLUMN_ALIASES = {
    "police_no": ["رقم الشرطة", "الرقم", "رقم", "police_no"],
    "name": ["الاسم", "اسم", "name"],
    "governorate": ["المحافظة", "محافظة", "governorate"],
    "qualification": ["المؤهل", "مؤهل", "qualification"],
    "batch": ["دفعة", "الدفعة", "batch"],
    "custody": ["العهدة", "عهدة", "custody"],
    "mahia": ["الماهية", "ماهية", "mahia", "تاريخ الماهية"],
    "phone": ["رقم الهاتف", "الهاتف", "موبايل", "تليفون", "phone"],
    "saraya": ["السرية", "سرية", "saraya"],
    "attend_date": ["تاريخ الحضور", "الحضور", "attend_date"],
    "religion": ["الديانة", "ديانة", "religion"],
    "notes": ["ملاحظات", "notes"],
    "craft": ["صنعة", "مهنة", "صنعة - مهنة", "craft"],
    "father_phone": ["تليفون الاب", "هاتف الأب", "father_phone"],
}

# الترتيب القياسي لأعمدة التصدير
EXPORT_COLUMNS = [
    ("م", "_row"),
    ("رقم الشرطة", "police_no"),
    ("الاسم", "name"),
    ("المحافظة", "governorate"),
    ("المؤهل", "qualification"),
    ("دفعة", "batch"),
    ("العهدة", "custody"),
    ("الماهية", "mahia"),
    ("رقم الهاتف", "phone"),
    ("الكتيبة", "battalion"),
    ("السرية", "saraya"),
    ("تاريخ الحضور", "attend_date"),
    ("العمالة", "employment"),
    ("الديانة", "religion"),
    ("الحالة", "status"),
    ("ملاحظات", "notes"),
]

STATUS_AR = {"present": "موجود", "leave": "إجازة", "absent": "غياب", "returned": "مُرحّل"}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _normalize(s: str) -> str:
    return str(s).strip().replace("ـ", "")


def _build_header_map(headers: List[str]) -> Dict[str, int]:
    """يرجّع: اسم الحقل -> رقم العمود (0-based)، من صف العناوين."""
    field_map: Dict[str, int] = {}
    norm = [_normalize(h) if h is not None else "" for h in headers]
    for field, aliases in COLUMN_ALIASES.items():
        for i, h in enumerate(norm):
            if h in [_normalize(a) for a in aliases]:
                field_map[field] = i
                break
    return field_map


def _read_rows(path: str) -> List[tuple]:
    """
    يقرأ كل صفوف أول ورقة من ملف Excel كقائمة tuples.
    يدعم .xlsx/.xlsm عبر openpyxl و .xls القديم عبر xlrd.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        try:
            import xlrd  # نوع الملف القديم (WPS / Excel 97-2003)
        except ImportError:
            raise ValueError(
                "لقراءة ملفات .xls القديمة ثبّت مكتبة xlrd:  pip install xlrd")
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        rows = []
        for r in range(sheet.nrows):
            vals = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                # تحويل تواريخ xlrd إلى نص YYYY-MM-DD
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        y, m, d, *_ = xlrd.xldate_as_tuple(cell.value, book.datemode)
                        vals.append(f"{y:04d}-{m:02d}-{d:02d}")
                        continue
                    except Exception:
                        pass
                v = cell.value
                # الأرقام الصحيحة بدون كسر عشري
                if isinstance(v, float) and v.is_integer():
                    v = int(v)
                vals.append(v)
            rows.append(tuple(vals))
        return rows
    # xlsx / xlsm
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def _find_header_row(rows: List[tuple], scan: int = 15) -> Tuple[int, Dict[str, int]]:
    """
    يبحث عن صف العناوين خلال أول `scan` صفوف (الملفات القديمة قد تبدأ بعناوين/شعار).
    يرجّع (فهرس صف العناوين، خريطة الحقول). يرفع ValueError لو لم يجد.
    """
    best_idx, best_map = -1, {}
    for idx in range(min(scan, len(rows))):
        hmap = _build_header_map(list(rows[idx]))
        if "police_no" in hmap and "name" in hmap:
            if len(hmap) > len(best_map):
                best_idx, best_map = idx, hmap
    if best_idx < 0:
        raise ValueError(
            "تعذّر العثور على عمودي 'رقم الشرطة' و'الاسم' في الملف. "
            "تأكد من صف العناوين.")
    return best_idx, best_map


def import_recruits_from_excel(path: str, intake_id: int, username: str = "import",
                               db_path: Optional[str] = None) -> Dict[str, Any]:
    """
    يقرأ ملف Excel (.xlsx / .xlsm / .xls) ويضيف المجندين.
    يرجّع: {added, skipped_duplicate, errors:[...], total_rows}
    """
    rows = _read_rows(path)
    if not rows:
        return {"added": 0, "skipped_duplicate": 0, "errors": [], "total_rows": 0}

    header_idx, header_map = _find_header_row(rows)
    rows = rows[header_idx:]  # تجاهل أي صفوف قبل العناوين

    added = skipped = 0
    errors: List[str] = []
    for ridx, row in enumerate(rows[1:], start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        data: Dict[str, Any] = {}
        for field, col in header_map.items():
            if col < len(row):
                val = row[col]
                data[field] = "" if val is None else str(val).strip()
        try:
            services.add_recruit(intake_id, data, username=username, db_path=db_path)
            added += 1
        except services.DuplicatePoliceNo:
            skipped += 1
        except Exception as e:  # noqa
            errors.append(f"صف {ridx}: {e}")
    return {"added": added, "skipped_duplicate": skipped,
            "errors": errors, "total_rows": len(rows) - 1}


def export_recruits_to_excel(intake_id: int, out_path: str,
                             title: str = "كشف المجندين",
                             db_path: Optional[str] = None) -> str:
    """يصدّر مجندي الدفعة إلى ملف Excel منسّق."""
    recruits = services.list_recruits(intake_id, db_path=db_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "المجندون"
    ws.sheet_view.rightToLeft = True

    # عنوان
    ncols = len(EXPORT_COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = Font(bold=True, size=14)
    tcell.alignment = _CENTER

    # العناوين
    for c, (header, _key) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=2, column=c, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER

    # البيانات
    for i, rec in enumerate(recruits, start=1):
        rrow = i + 2
        for c, (_header, key) in enumerate(EXPORT_COLUMNS, start=1):
            if key == "_row":
                val = i
            elif key == "status":
                val = STATUS_AR.get(rec.get("status"), rec.get("status"))
            else:
                val = rec.get(key, "")
            cell = ws.cell(row=rrow, column=c, value=val)
            cell.alignment = _CENTER
            cell.border = _BORDER

    # عرض الأعمدة
    widths = [5, 12, 24, 12, 11, 8, 10, 12, 14, 8, 7, 13, 9, 9, 9, 20]
    from openpyxl.utils import get_column_letter
    for c, w in enumerate(widths[:ncols], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return out_path


def generate_import_template(out_path: str) -> str:
    """ينشئ قالب استيراد فارغ بالأعمدة المطلوبة + مثال."""
    wb = Workbook()
    ws = wb.active
    ws.title = "قالب الاستيراد"
    ws.sheet_view.rightToLeft = True
    headers = ["رقم الشرطة", "الاسم", "المحافظة", "المؤهل", "دفعة", "العهدة",
               "الماهية", "رقم الهاتف", "السرية", "تاريخ الحضور", "الديانة",
               "صنعة - مهنة", "تليفون الاب", "ملاحظات"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
    # صف مثال
    example = ["123456", "محمد أحمد علي", "القاهرة", "متوسط", "1", "ميري",
               "2026-01-01", "01000000000", "1", "2026-01-01", "مسلم",
               "نجار", "01111111111", ""]
    for c, v in enumerate(example, start=1):
        ws.cell(row=2, column=c, value=v).alignment = _CENTER
    from openpyxl.utils import get_column_letter
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return out_path


def export_statistics_to_excel(stats: Dict[str, Any], out_path: str,
                               intake_name: str = "") -> str:
    """يصدّر البيان الإحصائي إلى ملف Excel منسّق."""
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "البيان"
    ws.sheet_view.rightToLeft = True

    ws.merge_cells("A1:B1")
    tcell = ws.cell(row=1, column=1, value=f"البيان الإحصائي — {intake_name}")
    tcell.font = Font(bold=True, size=14)
    tcell.alignment = _CENTER

    r = 3

    def _block(title, mapping, numeric=False):
        nonlocal r
        # عنوان القسم
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(row=r, column=1, value=title)
        c.fill = _HEADER_FILL; c.font = _HEADER_FONT; c.alignment = _CENTER
        c.border = _BORDER
        r += 1
        # رؤوس
        for col, h in enumerate(["البند", "العدد"], start=1):
            cell = ws.cell(row=r, column=col, value=h)
            cell.font = Font(bold=True); cell.alignment = _CENTER; cell.border = _BORDER
        r += 1
        items = list((mapping or {}).items())
        if numeric:
            items.sort(key=lambda x: (x[0] is None, x[0]))
        total = 0
        for k, v in items:
            label = "—" if k in (None, "", "None") else str(k)
            ws.cell(row=r, column=1, value=label).border = _BORDER
            vc = ws.cell(row=r, column=2, value=v); vc.border = _BORDER; vc.alignment = _CENTER
            try:
                total += int(v)
            except (ValueError, TypeError):
                pass
            r += 1
        tc1 = ws.cell(row=r, column=1, value="الإجمالي")
        tc1.font = Font(bold=True); tc1.border = _BORDER
        tc2 = ws.cell(row=r, column=2, value=total)
        tc2.font = Font(bold=True); tc2.border = _BORDER; tc2.alignment = _CENTER
        r += 2

    # الملخّص
    _block("ملخّص عام", {
        "الإجمالي": stats.get("total", 0),
        "الموجود": stats.get("existing", 0),
        "إجازات": stats.get("on_leave", 0),
        "غياب": stats.get("absent", 0),
        "حاضر": stats.get("present", 0),
    })
    _block("حسب المؤهل", stats.get("by_qualification", {}))
    _block("حسب العهدة", stats.get("by_custody", {}))
    _block("حسب الكتيبة", stats.get("by_battalion", {}), numeric=True)
    _block("حسب السرية", stats.get("by_saraya", {}), numeric=True)
    _block("حسب المحافظة", stats.get("by_governorate", {}))

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return out_path
